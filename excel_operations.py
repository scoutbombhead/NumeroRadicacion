"""
Excel file operations: reading and writing data
"""

from openpyxl import load_workbook
from config import EXCEL_FILE_PATH, logger


def read_numbers_from_excel():
    """
    Read all numbers from column D of the Excel file (excluding header)
    """
    try:
        logger.info(f"Reading numbers from Excel file: {EXCEL_FILE_PATH}")

        # Load the workbook
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        # Read all numbers from column D (4th column), skipping the header (first row)
        numbers = []
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True),
                                        start=2):
            if row[0] is not None:
                number = str(row[0]).strip()
                if number:
                    numbers.append(number)
                    logger.debug(f"Row {row_index}: {number}")

        if not numbers:
            logger.warning(f"No numbers found in Excel file")
            return []

        logger.info(f"Successfully read {len(numbers)} number(s) from Excel")
        return numbers

    except FileNotFoundError:
        logger.error(f"Excel file not found: {EXCEL_FILE_PATH}")
        return []
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
        return []


def write_data_to_excel(search_number, row_data):
    """
    Write the provided row_data (list of up to 6 values) into columns E-J (5-10)
    for the Excel row whose column D equals search_number.
    Returns True on success, False otherwise.
    """
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        target_row = None
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=4, max_col=4), start=2):
            cell_val = row[0].value
            if cell_val is not None and str(cell_val).strip() == str(search_number).strip():
                target_row = row_index
                break

        if target_row is None:
            logger.error(f"Number {search_number} not found in Excel file")
            return False

        # Ensure exactly 6 columns E-J
        values = row_data[:6]
        while len(values) < 6:
            values.append("")

        for i, val in enumerate(values):
            col = 5 + i  # Start from column E (5)
            cell = worksheet.cell(row=target_row, column=col)
            cell.value = val if val is not None else ""
            # Set cell to text format to avoid auto-formatting
            cell.number_format = '@'
            logger.debug(f"  Row {target_row}, Col {col}: {cell.value}")

        workbook.save(EXCEL_FILE_PATH)
        logger.info(f"Successfully wrote data to Excel for: {search_number}")
        return True
    except FileNotFoundError:
        logger.error(f"Excel file not found: {EXCEL_FILE_PATH}")
        return False
    except Exception as e:
        logger.error(f"Error writing to Excel: {str(e)}", exc_info=True)
        return False


def write_despacho_to_excel(search_number, despacho_value):
    """
    Write the despacho value into column C (3)
    for the Excel row whose column D equals search_number.
    Returns True on success, False otherwise.
    """
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        target_row = None
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=4, max_col=4), start=2):
            cell_val = row[0].value
            if cell_val is not None and str(cell_val).strip() == str(search_number).strip():
                target_row = row_index
                break

        if target_row is None:
            logger.error(f"Number {search_number} not found in Excel file")
            return False

        # Write despacho value to column C (3)
        cell = worksheet.cell(row=target_row, column=3)
        cell.value = despacho_value if despacho_value else ""
        # Set cell to text format to avoid auto-formatting
        cell.number_format = '@'
        logger.debug(f"  Row {target_row}, Col 3 (Despacho): {cell.value}")

        workbook.save(EXCEL_FILE_PATH)
        logger.info(f"Successfully wrote Despacho to Excel for: {search_number}")
        return True
    except FileNotFoundError:
        logger.error(f"Excel file not found: {EXCEL_FILE_PATH}")
        return False
    except Exception as e:
        logger.error(f"Error writing Despacho to Excel: {str(e)}", exc_info=True)
        return False


def write_sujetos_to_excel(search_number, demandante, demandado):
    """
    Write Demandante (column A) and Demandado (column B)
    for the Excel row whose column D equals search_number.
    Returns True on success, False otherwise.
    """
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        target_row = None
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=4, max_col=4), start=2):
            cell_val = row[0].value
            if cell_val is not None and str(cell_val).strip() == str(search_number).strip():
                target_row = row_index
                break

        if target_row is None:
            logger.error(f"Number {search_number} not found in Excel file")
            return False

        # Column A (1) -> Demandante, Column B (2) -> Demandado
        cell_a = worksheet.cell(row=target_row, column=1)
        cell_b = worksheet.cell(row=target_row, column=2)
        cell_a.value = demandante if demandante else ""
        cell_b.value = demandado if demandado else ""
        cell_a.number_format = '@'
        cell_b.number_format = '@'
        logger.debug(f"  Row {target_row}, Col 1 (Demandante): {cell_a.value}")
        logger.debug(f"  Row {target_row}, Col 2 (Demandado): {cell_b.value}")

        workbook.save(EXCEL_FILE_PATH)
        logger.info(f"Successfully wrote Demandante/Demandado to Excel for: {search_number}")
        return True
    except FileNotFoundError:
        logger.error(f"Excel file not found: {EXCEL_FILE_PATH}")
        return False
    except Exception as e:
        logger.error(f"Error writing Demandante/Demandado to Excel: {str(e)}", exc_info=True)
        return False
