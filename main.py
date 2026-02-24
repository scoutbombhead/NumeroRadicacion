from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
import re  # Used for timestamp pattern detection in table wait logic

# Global configuration
URL = "https://consultaprocesos.ramajudicial.gov.co/Procesos/NumeroRadicacion"
EXCEL_FILE_PATH = "C:\\Users\\sauda\\PycharmProjects\\NumeroRadicacion\\NumeroRadicacion.xlsx"  # Replace with your actual Excel file path


# ============================================================================
# EXCEL OPERATIONS
# ============================================================================

def read_numbers_from_excel():
    """
    Read all numbers from column D of the Excel file (excluding header)
    """
    try:
        print(f"Reading numbers from Excel file: {EXCEL_FILE_PATH}")

        # Load the workbook
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        # Read all numbers from column D (4th column), skipping the header (first row)
        numbers = []
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True),
                                        start=2):
            if row[0] is not None:
                number = str(row[0]).strip()
                if number:
                    numbers.append(number)
                    print(f"  Row {row_index}: {number}")

        if not numbers:
            print(f"✗ No numbers found in Excel file")
            return []

        print(f"✓ Successfully read {len(numbers)} number(s) from Excel")
        return numbers

    except FileNotFoundError:
        print(f"✗ Excel file not found: {EXCEL_FILE_PATH}")
        return []
    except Exception as e:
        print(f"✗ Error reading Excel file: {str(e)}")
        return []


# ============================================================================
# BROWSER OPERATIONS
# ============================================================================

def access_url():
    """
    Access the URL and print if successful
    """
    try:
        # Initialize the Chrome driver
        driver = webdriver.Chrome()

        print(f"Attempting to access: {URL}")
        driver.get(URL)

        # Wait a moment for the page to load
        time.sleep(2)

        # Check if page loaded successfully
        if driver.title:
            print(f"✓ Successfully accessed URL")
            print(f"✓ Page title: {driver.title}")
            return driver
        else:
            print("✗ Failed to load page (no title found)")
            driver.quit()
            return None

    except Exception as e:
        print(f"✗ Error accessing URL: {str(e)}")
        return None


# ============================================================================
# WEB INTERACTION - Page navigation and form filling
# ============================================================================

def select_second_radio_button(driver):
    """
    Select the 2nd radio button on the page
    """
    try:
        # Wait for radio buttons to be present
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "input[type='radio']"))
        )

        # Find all radio buttons
        radio_buttons = driver.find_elements(By.CSS_SELECTOR, "input[type='radio']")

        if len(radio_buttons) < 2:
            print(f"✗ Found only {len(radio_buttons)} radio button(s), need at least 2")
            return False

        # Use JavaScript to click the 2nd radio button to bypass click interception
        second_radio = radio_buttons[1]
        driver.execute_script("arguments[0].click();", second_radio)
        print(f"✓ Successfully selected 2nd radio button")
        time.sleep(1)
        return True

    except Exception as e:
        print(f"✗ Error selecting radio button: {str(e)}")
        return False


def enter_search_number(driver, search_number):
    """
    Enter the search number into the text field
    """
    try:
        # Wait for the text input field to be present
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "input[type='text']"))
        )

        # Find the text input field
        text_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='text']")

        if len(text_inputs) == 0:
            print(f"✗ No text input field found")
            return False

        # Use the first text input field (most likely the search bar)
        search_field = text_inputs[0]

        # Clear any existing value and enter the search number
        search_field.clear()
        search_field.send_keys(search_number)
        print(f"✓ Successfully entered search number: {search_number}")
        time.sleep(1)
        return True

    except Exception as e:
        print(f"✗ Error entering search number: {str(e)}")
        return False


def click_consultar_button(driver):
    """
    Click the CONSULTAR button - tries multiple strategies to find it
    """
    try:
        # Strategy 1: Find button by aria-label containing "Consultar"
        try:
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'Consultar')]"))
            )
            consultar_button = driver.find_element(By.XPATH, "//button[contains(@aria-label, 'Consultar')]")
            driver.execute_script("arguments[0].click();", consultar_button)
            print(f"✓ Successfully clicked CONSULTAR button (Strategy 1 - aria-label)")
            time.sleep(2)
            return True
        except:
            pass

        # Strategy 2: Find button by span text inside button
        try:
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button//span[contains(text(), 'Consultar')]/.."))
            )
            consultar_button = driver.find_element(By.XPATH, "//button//span[contains(text(), 'Consultar')]/..")
            driver.execute_script("arguments[0].click();", consultar_button)
            print(f"✓ Successfully clicked CONSULTAR button (Strategy 2 - span text)")
            time.sleep(2)
            return True
        except:
            pass

        # Strategy 3: Find by button with success class
        try:
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'success')]//span[contains(text(), 'Consultar')]"))
            )
            consultar_button = driver.find_element(By.XPATH,
                                                   "//button[contains(@class, 'success')]//span[contains(text(), 'Consultar')]/../..")
            driver.execute_script("arguments[0].click();", consultar_button)
            print(f"✓ Successfully clicked CONSULTAR button (Strategy 3 - success class)")
            time.sleep(2)
            return True
        except:
            pass

        print(f"✗ Could not find CONSULTAR button with any strategy")
        return False

    except Exception as e:
        print(f"✗ Error clicking CONSULTAR button: {str(e)}")
        return False


def click_volver_button(driver):
    """
    Click the VOLVER button if it appears (in dialog)
    This button may not always appear, so we use a shorter timeout
    """
    try:
        # Strategy 1: Find button by span text "Volver" inside button
        try:
            WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//button//span[contains(text(), 'Volver')]"))
            )
            volver_button = driver.find_element(By.XPATH, "//button//span[contains(text(), 'Volver')]/..")
            driver.execute_script("arguments[0].click();", volver_button)
            print(f"✓ Successfully clicked VOLVER button (Strategy 1 - span text)")
            time.sleep(1)
            return True
        except:
            pass

        # Strategy 2: Find button with aria-label containing "Volver"
        try:
            WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'Volver')]"))
            )
            volver_button = driver.find_element(By.XPATH, "//button[contains(@aria-label, 'Volver')]")
            driver.execute_script("arguments[0].click();", volver_button)
            print(f"✓ Successfully clicked VOLVER button (Strategy 2 - aria-label)")
            time.sleep(1)
            return True
        except:
            pass

        # If no VOLVER button is found, it's likely not needed for this search
        print(f"✓ No VOLVER button found (dialog may not have appeared)")
        return True

    except Exception as e:
        print(f"✗ Error clicking VOLVER button: {str(e)}")
        return False


def click_first_clickable_table_number(driver):
    """
    Find the table and click the first clickable number in the second column
    Starts from the second row (after headers) and checks each row
    """
    try:
        # Wait for the table to be present
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
        )

        # Find the table
        table = driver.find_element(By.TAG_NAME, "table")

        # Get all rows in the table body
        rows = table.find_elements(By.TAG_NAME, "tr")

        if len(rows) < 2:
            print(f"✗ Table has less than 2 rows (no data rows)")
            return False

        # Iterate through rows starting from the second row (index 1) to skip headers
        for index, row in enumerate(rows[1:], start=2):
            try:
                # Get all cells in this row
                cells = row.find_elements(By.TAG_NAME, "td")

                if len(cells) < 2:
                    print(f"✗ Row {index} has less than 2 columns, skipping")
                    continue

                # Get the second column (index 1)
                second_column = cells[1]

                # Check if there's a clickable button in the second column
                try:
                    button = second_column.find_element(By.TAG_NAME, "button")
                    # If we found a button, it's clickable
                    number_text = button.text.strip()
                    print(f"✓ Found clickable number in row {index}: {number_text}")

                    # Click the button
                    driver.execute_script("arguments[0].click();", button)
                    print(f"✓ Successfully clicked the number: {number_text}")
                    time.sleep(2)
                    return True

                except:
                    # No button found in this row, check for text
                    try:
                        paragraph = second_column.find_element(By.TAG_NAME, "p")
                        number_text = paragraph.text.strip()
                        print(f"✓ Row {index} has non-clickable number: {number_text}")
                    except:
                        pass
                    # Continue to the next row
                    continue

            except Exception as e:
                print(f"✗ Error processing row {index}: {str(e)}")
                continue

        print(f"✗ No clickable numbers found in table")
        return False

    except Exception as e:
        print(f"✗ Error accessing table: {str(e)}")
        return False


def click_actuaciones_tab(driver):
    """
    Click the Actuaciones tab
    """
    try:
        # Wait for the tab to be present
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//div[@role='tab']"))
        )

        # Find the Actuaciones tab by searching for a div with role="tab" containing "Actuaciones"
        actuaciones_tab = driver.find_element(By.XPATH, "//div[@role='tab'][contains(text(), 'Actuaciones')]")

        # Click the tab using JavaScript to avoid any click interception
        driver.execute_script("arguments[0].click();", actuaciones_tab)
        print(f"✓ Successfully clicked Actuaciones tab")
        # Wait a short moment for tab content to load
        time.sleep(1)

        # Wait until an actuaciones table appears to be ready (has a non-timestamp cell)
        if wait_for_actuaciones_table_ready(driver, timeout=8):
            return True
        else:
            # Still return True; printing/debugging will help identify issues
            print("✱ Warning: Actuaciones table may not be fully loaded yet")
            return True

    except Exception as e:
        print(f"✗ Error clicking Actuaciones tab: {str(e)}")
        return False


# ============================================================================
# TABLE EXTRACTION - Extract data from Actuaciones table
# ============================================================================

def wait_for_actuaciones_table_ready(driver, timeout=8):
    """
    Wait until at least one table row contains a non-empty, non-timestamp cell.
    Returns True if ready, False on timeout.
    """
    end_time = time.time() + timeout
    timestamp_re = re.compile(r"\d{4}-\d{2}-\d{2}|\d{1,2}\s+[A-Za-z]{3,}\s+\d{4}|\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}")
    try:
        while time.time() < end_time:
            tables = driver.find_elements(By.TAG_NAME, "table")
            for t in tables:
                rows = t.find_elements(By.TAG_NAME, "tr")
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if not cells:
                        cells = row.find_elements(By.TAG_NAME, "th")
                    for cell in cells:
                        try:
                            text = (cell.get_attribute('innerText') or cell.text or "").strip()
                        except:
                            text = (cell.text or "").strip()
                        if not text:
                            continue
                        # if the text doesn't look like a timestamp, consider table ready
                        if not timestamp_re.search(text):
                            return True
            time.sleep(0.5)
        return False
    except Exception:
        return False


def print_actuaciones_first_row(driver):
    """
    Find the Actuaciones table (using the same heuristics) and print only the first row.
    Returns the list of cell values for the first row (padded to 6 columns).
    """
    try:
        # Wait for tables to appear
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
        )

        tables = driver.find_elements(By.TAG_NAME, "table")
        if not tables:
            print("✗ No tables found on Actuaciones tab")
            return []

        # Reuse the same selection heuristics as print_actuaciones_table
        chosen_table = None
        best_col_count = 0
        header_keywords = ["fecha", "actuaci", "despacho", "tipo", "documento", "observacion"]
        for t in tables:
            rows = t.find_elements(By.TAG_NAME, "tr")
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "th")
                if not cells:
                    cells = row.find_elements(By.TAG_NAME, "td")
                if cells:
                    try:
                        header_text = " ".join(
                            [(cell.get_attribute('innerText') or cell.text).strip().lower() for cell in cells])
                    except Exception:
                        header_text = ""

                    if any(k in header_text for k in header_keywords):
                        chosen_table = t
                        best_col_count = len(cells)
                        break

                    col_count = len(cells)
                    if col_count == 6:
                        chosen_table = t
                        best_col_count = col_count
                        break
                    if col_count > best_col_count:
                        chosen_table = t
                        best_col_count = col_count
                    break
            if best_col_count == 6:
                break

        if chosen_table is None:
            chosen_table = tables[0]

        # Prefer rows in tbody if present (these are typically data rows)
        tbodies = chosen_table.find_elements(By.TAG_NAME, "tbody")
        if tbodies:
            rows = tbodies[0].find_elements(By.TAG_NAME, "tr")
        else:
            rows = chosen_table.find_elements(By.TAG_NAME, "tr")

        if not rows:
            print("✗ Actuaciones table has no rows")
            return []

        # Determine the first data row (row below header):
        # - If tbody was used, rows[0] is likely the first data row.
        # - Otherwise, if the table has multiple rows, assume rows[0] is header and use rows[1].
        if tbodies:
            data_row = rows[0]
        else:
            data_row = rows[1] if len(rows) > 1 else rows[0]
        cells = data_row.find_elements(By.TAG_NAME, "td")
        if not cells:
            cells = data_row.find_elements(By.TAG_NAME, "th")

        row_values = []
        for c in range(6):
            if c < len(cells):
                cell = cells[c]
                try:
                    # prefer common nested elements
                    try:
                        el = cell.find_element(By.TAG_NAME, "button")
                        text = el.get_attribute('innerText') or el.text
                    except:
                        try:
                            el = cell.find_element(By.TAG_NAME, "span")
                            text = el.get_attribute('innerText') or el.text
                        except:
                            try:
                                el = cell.find_element(By.TAG_NAME, "a")
                                text = el.get_attribute('innerText') or el.text
                            except:
                                text = cell.get_attribute('innerText') or cell.text
                except Exception:
                    text = cell.get_attribute('innerText') or cell.text

                if text is None:
                    text = ""
                row_values.append(text.strip())
            else:
                row_values.append("")

        print(f"Actuaciones first row: {row_values}")
        return row_values

    except Exception as e:
        print(f"✗ Error printing first row of Actuaciones table: {str(e)}")
        return []


# ============================================================================
# DATOS DE PROCESO - Extract data from Datos de Proceso tab
# ============================================================================

# ============================================================================
# DATOS DE PROCESO - Extract data from Datos de Proceso tab
# ============================================================================

def extract_despacho_value(driver):
    """
    Extract the "Despacho" value from the Datos de Proceso table
    Returns the despacho value as a string, or empty string if not found
    """
    try:
        # Wait for the table to be present
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
        )

        # Get all tables on the page
        tables = driver.find_elements(By.TAG_NAME, "table")

        if not tables:
            print(f"✗ No tables found in Datos de Proceso")
            return ""

        # Try each table to find the one with Despacho
        for table_idx, table in enumerate(tables):
            rows = table.find_elements(By.TAG_NAME, "tr")

            if not rows:
                continue

            # Look for the row that contains "Despacho:"
            for row in rows:
                try:
                    # Get th and td elements separately (preserve order in DOM)
                    th_elements = row.find_elements(By.TAG_NAME, "th")
                    td_elements = row.find_elements(By.TAG_NAME, "td")

                    if not th_elements or not td_elements:
                        continue

                    # Check if first th is "Despacho:"
                    first_th_text = (th_elements[0].get_attribute('innerText') or th_elements[0].text or "").strip()

                    if first_th_text == "Despacho:" and len(td_elements) > 0:
                        # Get the value from the first td in this row
                        despacho_value = (
                                    td_elements[0].get_attribute('innerText') or td_elements[0].text or "").strip()

                        # Skip if it's a date (YYYY-MM-DD format) or if it looks like a label (contains ":")
                        is_date = len(despacho_value) == 10 and despacho_value[4] == '-' and despacho_value[7] == '-'
                        is_label = ":" in despacho_value and len(despacho_value) < 30

                        if despacho_value and not is_date and not is_label:
                            print(f"✓ Found Despacho: {despacho_value}")
                            return despacho_value
                except:
                    continue

        # If we get here, we didn't find it - print debug info
        print(f"✗ Despacho row not found in table")
        print_datos_proceso_table_debug(driver)
        return ""

    except Exception as e:
        print(f"✗ Error extracting Despacho value: {str(e)}")
        return ""


def print_datos_proceso_table_debug(driver):
    """
    Debug helper: print all table rows to help identify the structure
    """
    try:
        tables = driver.find_elements(By.TAG_NAME, "table")
        print(f"\nDebug: Found {len(tables)} table(s) on Datos de Proceso tab")

        for t_idx, table in enumerate(tables):
            rows = table.find_elements(By.TAG_NAME, "tr")
            print(f"  Table {t_idx + 1}: {len(rows)} row(s)")

            for r_idx, row in enumerate(rows[:10]):  # Print first 10 rows
                try:
                    # Get all cells
                    cells = row.find_elements(By.TAG_NAME, "th")
                    if not cells:
                        cells = row.find_elements(By.TAG_NAME, "td")

                    row_data = []
                    for cell in cells:
                        text = (cell.get_attribute('innerText') or cell.text or "").strip()
                        row_data.append(
                            text[:50] + ("..." if len(text) > 50 else ""))  # Limit to 50 chars for readability

                    print(f"    Row {r_idx + 1}: {row_data}")
                except Exception as e:
                    print(f"    Row {r_idx + 1}: [Error reading row]")
    except Exception as e:
        print(f"✗ Error printing debug info: {str(e)}")


# ============================================================================
# SUBJETOS PROCESALES - Extract data from Subjetos Procesales tab
# ============================================================================

def click_subjetos_procesales_tab(driver):
    """
    Click the Sujetos Procesales tab
    """
    try:
        # Wait for the tab to be present
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//div[@role='tab']"))
        )

        # Find the Sujetos Procesales tab by searching for a div with role="tab" containing "Sujetos Procesales"
        tab = driver.find_element(By.XPATH, "//div[@role='tab'][contains(text(), 'Sujetos Procesales')]")

        # Click the tab using JavaScript to avoid any click interception
        driver.execute_script("arguments[0].click();", tab)
        print(f"✓ Successfully clicked Sujetos Procesales tab")
        time.sleep(2)
        return True

    except Exception as e:
        print(f"✗ Error clicking Sujetos Procesales tab: {str(e)}")
        return False


def extract_subjetos_procesales(driver):
    """
    Extract Demandante and Demandado values from the Subjetos Procesales table
    Returns a dictionary with 'demandante' and 'demandado' keys
    """
    try:
        # Wait for the table to be present
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
        )

        # Get all tables on the page
        tables = driver.find_elements(By.TAG_NAME, "table")

        if not tables:
            print(f"✗ No tables found in Subjetos Procesales")
            return {"demandante": "", "demandado": ""}

        demandante = ""
        demandado = ""

        # Try each table to find the required data
        for table_idx, table in enumerate(tables):
            rows = table.find_elements(By.TAG_NAME, "tr")

            if not rows:
                continue

            # Look for rows containing "Demandante" and "Demandando"
            for row in rows:
                try:
                    # Get all cells in this row
                    cells = row.find_elements(By.TAG_NAME, "td")

                    if len(cells) >= 2:
                        first_cell_text = (cells[0].get_attribute('innerText') or cells[0].text or "").strip()
                        second_cell_text = (cells[1].get_attribute('innerText') or cells[1].text or "").strip()

                        # Check for Demandante
                        if first_cell_text == "Demandante" and not demandante:
                            demandante = second_cell_text
                            print(f"✓ Found Demandante: {demandante}")

                        # Check for Demandado
                        if first_cell_text == "Demandado" and not demandado:
                            demandado = second_cell_text
                            print(f"✓ Found Demandado: {demandado}")

                        # If we have both, we can stop searching
                        if demandante and demandado:
                            return {"demandante": demandante, "demandado": demandado}

                except:
                    continue

        print(f"Demandante: {demandante if demandante else '(not found)'}")
        print(f"Demandado: {demandado if demandado else '(not found)'}")
        return {"demandante": demandante, "demandado": demandado}

    except Exception as e:
        print(f"✗ Error extracting Subjetos Procesales data: {str(e)}")
        return {"demandante": "", "demandado": ""}


# ============================================================================
# DATA WRITING - Write extracted data to Excel
# ============================================================================

def write_data_to_excel(search_number, row_data):
    """
    Write the provided row_data (list of up to 6 values) into columns E-J (5-10)
    for the Excel row whose column D equals search_number.
    Returns True on success, False otherwise.
    """
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        target_row = None
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=4, max_col=4), start=2):
            cell_val = row[0].value
            if cell_val is not None and str(cell_val).strip() == str(search_number).strip():
                target_row = row_index
                break

        if target_row is None:
            print(f"✗ Number {search_number} not found in Excel file")
            return False

        # Ensure exactly 6 columns E-J
        values = row_data[:6]
        while len(values) < 6:
            values.append("")

        for i, val in enumerate(values):
            col = 5 + i  # Start from column E (5)
            cell = worksheet.cell(row=target_row, column=col)
            cell.value = val if val is not None else ""
            # Set cell to text format to avoid auto-formatting
            cell.number_format = '@'
            print(f"  Row {target_row}, Col {col}: {cell.value}")

        workbook.save(EXCEL_FILE_PATH)
        print(f"✓ Successfully wrote data to Excel for: {search_number}")
        return True
    except FileNotFoundError:
        print(f"✗ Excel file not found: {EXCEL_FILE_PATH}")
        return False
    except Exception as e:
        print(f"✗ Error writing to Excel: {str(e)}")
        return False


def write_despacho_to_excel(search_number, despacho_value):
    """
    Write the despacho value into column C (3)
    for the Excel row whose column D equals search_number.
    Returns True on success, False otherwise.
    """
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        target_row = None
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=4, max_col=4), start=2):
            cell_val = row[0].value
            if cell_val is not None and str(cell_val).strip() == str(search_number).strip():
                target_row = row_index
                break

        if target_row is None:
            print(f"✗ Number {search_number} not found in Excel file")
            return False

        # Write despacho value to column C (3)
        cell = worksheet.cell(row=target_row, column=3)
        cell.value = despacho_value if despacho_value else ""
        # Set cell to text format to avoid auto-formatting
        cell.number_format = '@'
        print(f"  Row {target_row}, Col 3 (Despacho): {cell.value}")

        workbook.save(EXCEL_FILE_PATH)
        print(f"✓ Successfully wrote Despacho to Excel for: {search_number}")
        return True
    except FileNotFoundError:
        print(f"✗ Excel file not found: {EXCEL_FILE_PATH}")
        return False
    except Exception as e:
        print(f"✗ Error writing Despacho to Excel: {str(e)}")
        return False


def write_sujetos_to_excel(search_number, demandante, demandado):
    """
    Write Demandante (column A) and Demandado (column B)
    for the Excel row whose column D equals search_number.
    Returns True on success, False otherwise.
    """
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        target_row = None
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=4, max_col=4), start=2):
            cell_val = row[0].value
            if cell_val is not None and str(cell_val).strip() == str(search_number).strip():
                target_row = row_index
                break

        if target_row is None:
            print(f"✗ Number {search_number} not found in Excel file")
            return False

        # Column A (1) -> Demandante, Column B (2) -> Demandado
        cell_a = worksheet.cell(row=target_row, column=1)
        cell_b = worksheet.cell(row=target_row, column=2)
        cell_a.value = demandante if demandante else ""
        cell_b.value = demandado if demandado else ""
        cell_a.number_format = '@'
        cell_b.number_format = '@'
        print(f"  Row {target_row}, Col 1 (Demandante): {cell_a.value}")
        print(f"  Row {target_row}, Col 2 (Demandado): {cell_b.value}")

        workbook.save(EXCEL_FILE_PATH)
        print(f"✓ Successfully wrote Demandante/Demandado to Excel for: {search_number}")
        return True
    except FileNotFoundError:
        print(f"✗ Excel file not found: {EXCEL_FILE_PATH}")
        return False
    except Exception as e:
        print(f"✗ Error writing Demandante/Demandado to Excel: {str(e)}")
        return False


# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == '__main__':
    # Read all numbers from Excel file
    search_numbers = read_numbers_from_excel()

    if not search_numbers:
        print("No numbers to search. Exiting.")
    else:
        # Loop through each number
        for index, number in enumerate(search_numbers, start=1):
            print(f"\n{'=' * 60}")
            print(f"Processing number {index}/{len(search_numbers)}: {number}")
            print(f"{'=' * 60}")

            # Start by accessing the URL
            driver = access_url()

            if driver:
                try:
                    # Select the 2nd radio button
                    select_second_radio_button(driver)

                    # Enter the search number
                    enter_search_number(driver, number)

                    # Click the CONSULTAR button
                    click_consultar_button(driver)

                    # Click the VOLVER button if a dialog appears
                    click_volver_button(driver)

                    # Click the first clickable number in the table's second column
                    click_first_clickable_table_number(driver)

                    # Extract the Despacho value from the Datos de Proceso tab (default view)
                    despacho_value = extract_despacho_value(driver)
                    if despacho_value:
                        print(f"✓ Extracted Despacho value: {despacho_value}")
                        # Write Despacho to column C
                        write_despacho_to_excel(number, despacho_value)
                    else:
                        print(f"✗ Failed to extract Despacho value for {number}")

                    # Click the Actuaciones tab
                    click_actuaciones_tab(driver)

                    # Get the first data row from Actuaciones and write it to Excel (columns B-G)
                    first_row = print_actuaciones_first_row(driver)
                    if first_row:
                        # Ensure exactly 6 columns (B-G)
                        row_values = first_row[:6]
                        while len(row_values) < 6:
                            row_values.append("")

                        print(f"✓ Actuaciones first row has {len(row_values)} columns: {row_values}")
                        written = write_data_to_excel(number, row_values)
                        if written:
                            print(f"✓ Wrote Actuaciones first row to Excel for {number}")
                        else:
                            print(f"✗ Failed to write Actuaciones data to Excel for {number}")
                    else:
                        print(f"✓ No first row found in Actuaciones table")

                    # Write the Despacho value to Excel (column C)
                    if despacho_value:
                        write_despacho_to_excel(number, despacho_value)

                    # Click the Subjetos Procesales tab
                    click_subjetos_procesales_tab(driver)

                    # Extract Demandante and Demandando values
                    sujetos_data = extract_subjetos_procesales(driver)
                    if sujetos_data["demandante"] or sujetos_data["demandado"]:
                        print(f"✓ Extracted Subjetos Procesales data:")
                        print(f"  Demandante: {sujetos_data['demandante']}")
                        print(f"  Demandado: {sujetos_data['demandado']}")
                        write_sujetos_to_excel(number, sujetos_data["demandante"], sujetos_data["demandado"])
                    else:
                        print(f"✗ Failed to extract Subjetos Procesales data for {number}")

                    print(f"✓ Successfully completed search for: {number}")

                except Exception as e:
                    print(f"✗ Error during search for {number}: {str(e)}")

                finally:
                    # Close the driver after each search
                    driver.quit()
            else:
                print(f"Failed to initialize browser for: {number}")

            # Add a small delay between searches
            if index < len(search_numbers):
                print(f"Waiting before next search...")
                time.sleep(2)

        print(f"\n{'=' * 60}")
        print(f"✓ All {len(search_numbers)} searches completed!")
        print(f"{'=' * 60}")